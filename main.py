import asyncio
import configparser
from msgraph.generated.models.o_data_errors.o_data_error import ODataError
from graph import Graph
from msgraph import GraphServiceClient
#from msgraph.generated.graph_client import GraphClient
from azure.identity import DeviceCodeCredential
from msgraph.graph_service_client import GraphServiceClient
#from msgraph.generated.me.calendar.events.events_request_builder import EventsRequestBuilderGetRequestConfiguration
from msgraph.generated.models.o_data_errors.o_data_error import ODataError
from msgraph import GraphServiceClient
from msgraph.generated.users.item.calendar.events.events_request_builder import EventsRequestBuilder
from kiota_abstractions.base_request_configuration import RequestConfiguration
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
from openpyxl.styles import NamedStyle
from datetime import datetime
# Removed unused import that could not be resolved
#from msgraph.core import GraphClient



#result = await graph_client.me.calendar.events.get()
async def main():
    scopes = ['User.Read', 'Calendars.Read']

# Multi-tenant apps can use "common",
# single-tenant apps must use the tenant ID from the Azure portal
    tenant_id = '94aa9436-2653-434c-bd47-1124432cb7d7'

# Values from app registration
    client_id = 'b9d8662f-5800-4498-a4f1-28b92bea4f39'

# azure.identity
    credential = DeviceCodeCredential(
        tenant_id=tenant_id,
        client_id=client_id)
    graph_client = GraphServiceClient(credential, scopes)
    query_params = EventsRequestBuilder.EventsRequestBuilderGetQueryParameters(
		filter="contains(subject,'GO')",
    )

    request_configuration = RequestConfiguration(
    query_parameters = query_params,
    )

    result = await graph_client.me.calendar.events.get(request_configuration = request_configuration)
    #print(result)


    excel_file = "GO.xlsx"
    sheet_name = "Sheet1"

    filtered_events = []
    for event in result.value:
        if re.search(r'\bGO\b', event.subject, re.IGNORECASE):
            print(f"{event.subject} at {event.start.date_time} {event.end.date_time}")
            filtered_events.append({
                'subject': event.subject,
                'start': event.start.date_time,
                'end': event.end.date_time
            })
    # Create or open the workbook
    if os.path.exists(excel_file): 
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Write header if creating new file
        ws.append(["Subject", "Start", "End"])
        # Optional: bold headers
        for col in range(1, 4):
            ws.cell(row=1, column=col).font = Font(bold=True)

    # if "date_style" not in wb.named_styles:
    #     date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
    #     wb.add_named_style(date_style)
    # else:
    #     date_style = wb.named_styles["date_style"]

   #date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.value = None
    # Find next empty row
    next_row = 2

    # Append new events
    for i, event_data in enumerate(filtered_events, start=next_row):
        start_time = datetime.fromisoformat(event_data['start'])
        end_time = datetime.fromisoformat(event_data['end'])
    
        ws.cell(row=i, column=1, value=event_data['subject'])
        ws.cell(row=i, column=2, value=start_time)
        ws.cell(row=i, column=3, value=end_time)

    # Save
    wb.save(excel_file)
    print(f"Appended {len(filtered_events)} events to {excel_file}")


asyncio.run(main())


